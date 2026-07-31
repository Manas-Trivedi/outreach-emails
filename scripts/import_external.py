# --- path shim: resolve data paths against ../industry regardless of CWD ---
import os as _os
_os.chdir(_os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "..", "industry"))
# --- end shim ---
import csv
import os

# Imports external contact files into extra_contacts.csv (same schema as hr_contacts.csv).
# gen_batches.py merges extra_contacts.csv into the sendable batch files.

SRC = r"D:\New folder\python-3.12.3-amd64"

LINKEDIN_CSVS = [
    "jaipur.csv",
    "Multiple_Folders_2026-04-06 (1).csv",
    "Multiple_Folders_2026-04-06.csv",
]
GROWTH_XLSX = "growth-list-free-forever-march-2026.xlsx"  # (1) copy is identical

out = []
seen = set()

# Cumulative: keep everything already imported (source files may get moved/deleted)
if os.path.exists("extra_contacts.csv"):
    with open("extra_contacts.csv", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            k = r["Email"].lower()
            if k not in seen:
                seen.add(k)
                out.append([r["Company"], r["Name"], r["Title"], r["Email"], r["LinkedIn"], r["Action"]])

def add(company, name, title, email, linkedin, notes):
    email = (email or "").strip()
    if "@" not in email or " " in email:
        return
    key = email.lower()
    if key in seen:
        return
    seen.add(key)
    out.append([company or "—", name or "—", title or "—", email, linkedin or "", notes])

# LinkedIn people exports
for fname in LINKEDIN_CSVS:
    path = os.path.join(SRC, fname)
    if not os.path.exists(path):
        print("MISSING:", path)
        continue
    with open(path, encoding="utf-8", errors="replace") as f:
        for r in csv.DictReader(f):
            name = " ".join(x for x in [r.get("First Name", ""), r.get("Last Name", "")] if x).strip()
            work = r.get("Work Email", "").strip()
            personal = r.get("Personal Email", "").strip()
            extras = [r.get(k, "").strip() for k in ("Additional Email 1", "Additional Email 2", "Additional Email 3")]
            emails = [e for e in [work, personal] + extras if "@" in e]
            if not emails:
                continue
            primary, alts = emails[0], emails[1:]
            notes_bits = []
            if alts:
                notes_bits.append("alt: " + ", ".join(alts))
            for pk in ("Phone", "Phone 2"):
                if r.get(pk, "").strip():
                    notes_bits.append("Ph: " + r[pk].strip())
            loc = r.get("Location", "").split(",")[0].strip()
            if loc:
                notes_bits.append(loc)
            add(r.get("Company", ""), name, r.get("Title", ""), primary,
                r.get("Linkedin URL", ""), "; ".join(notes_bits))

# Growth List funded companies (verified contact emails)
from openpyxl import load_workbook
gpath = os.path.join(SRC, GROWTH_XLSX)
if os.path.exists(gpath):
    ws = load_workbook(gpath, read_only=True).worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    hdr = {str(v): i for i, v in enumerate(rows[0])}
    for r in rows[1:]:
        if not r or not r[hdr["Contact Email"]]:
            continue
        status = str(r[hdr["Email Status"]] or "")
        funding = str(r[hdr["Funding Amount (in USD)"]] or "")
        notes = "; ".join(x for x in [
            "Growth List Mar-2026 funded co",
            f"{r[hdr['City']]}, {r[hdr['Country']]}",
            f"raised ${funding}" if funding not in ("", "unknown", "None") else "",
            str(r[hdr["Funding Type"]] or ""),
            f"email {status}",
        ] if x)
        add(str(r[hdr["Name"]] or ""), "Team", "Contact", str(r[hdr["Contact Email"]]),
            str(r[hdr["LinkedIn"]] or ""), notes)

# ── Batch 2 of external files ─────────────────────────────────────────────────

def sheet_rows(fname):
    path = os.path.join(SRC, fname)
    if not os.path.exists(path):
        print("MISSING:", path)
        return []
    ws = load_workbook(path, read_only=True).worksheets[0]
    return [[("" if c is None else str(c).strip()) for c in row] for row in ws.iter_rows(values_only=True)]

# Angel investors (no header; col0=name, col1=emails possibly newline-separated)
for r in sheet_rows("912747609-Angels-Emails.xlsx"):
    if len(r) < 2 or "@" not in r[1]:
        continue
    emails = [e.strip() for e in r[1].replace("\n", " ").replace(",", " ").split() if "@" in e]
    if not emails:
        continue
    notes = "Angel investor list" + ("; alt: " + ", ".join(emails[1:]) if len(emails) > 1 else "")
    add("(Angel Investor)", r[0], "Angel Investor", emails[0], "", notes)

# andy emails.xlsx — Company, Email
for r in sheet_rows("andy emails.xlsx")[1:]:
    if len(r) >= 2:
        add(r[0], "Team", "Contact", r[1], "", "andy company-emails list")

# Be.xlsx — Name, Company, Email (1.3K Bengaluru contacts)
for r in sheet_rows("Be.xlsx")[1:]:
    if len(r) >= 3:
        add(r[1], r[0], "—", r[2], "", "Bengaluru contact list (Be.xlsx)")

# Bengaluru.xlsx — name, email, company, company name
for r in sheet_rows("Bengaluru.xlsx")[1:]:
    if len(r) >= 2 and "@" in r[1]:
        name = r[0] if r[0] and "..." not in r[0] else r[1].split("@")[0]
        company = (r[2] if len(r) > 2 and r[2] else (r[3] if len(r) > 3 else "")) or r[1].split("@")[1].split(".")[0]
        add(company, name, "—", r[1], "", "Bengaluru list")

# emails_companies.xlsx — Company, Email (loosely structured)
for r in sheet_rows("emails_companies.xlsx")[1:]:
    if len(r) >= 2:
        add(r[0], r[0], "—", r[1], "", "emails_companies list")

# final_contacts.xlsx — Name, Company, Email, Role
for r in sheet_rows("final_contacts.xlsx")[1:]:
    if len(r) >= 3:
        name = r[0] if r[0] and "..." not in r[0] else (r[2].split("@")[0] if "@" in r[2] else r[0])
        add(r[1], name, r[3] if len(r) > 3 else "—", r[2], "", "final_contacts list")

# final_contacts (1)/(2).xlsx — Name, Company, Email (subsets, dedup handles overlap)
for fn in ["final_contacts (1).xlsx", "final_contacts (2).xlsx"]:
    for r in sheet_rows(fn)[1:]:
        if len(r) >= 3:
            add(r[1], r[0], "—", r[2], "", "final_contacts subset")

# Bengaluru tech recruiters CSV
rpath = os.path.join(SRC, "bengaluru_tech_recruiters_emails_20260404161938_core.csv")
if os.path.exists(rpath):
    with open(rpath, encoding="utf-8", errors="replace") as f:
        for r in csv.DictReader(f):
            email = r.get("contact_professions_email", "").strip()
            if "@" not in email:
                raw = r.get("contact_emails", "")
                for tok in raw.replace('"', " ").replace(":", " ").replace(",", " ").split():
                    if "@" in tok and "." in tok:
                        email = tok.strip("[]{}() ")
                        break
            notes_bits = ["Bengaluru tech recruiter"]
            if r.get("contact_mobile_phone", "").strip():
                notes_bits.append("Ph: " + r["contact_mobile_phone"].strip())
            add(r.get("prospect_company_name", ""), r.get("prospect_full_name", ""),
                r.get("prospect_job_title", "")[:80], email,
                r.get("prospect_linkedin", ""), "; ".join(notes_bits))

# Myntra leadership export (ZoomInfo-style)
mpath = os.path.join(SRC, "698b7bd1e1a802156d792ce1_1770749556922.csv")
if os.path.exists(mpath):
    with open(mpath, encoding="utf-8", errors="replace") as f:
        for r in csv.DictReader(f):
            name = " ".join(x for x in [r.get("First Name", ""), r.get("Last Name", "")] if x).strip()
            notes_bits = ["Myntra leadership list"]
            if r.get("Email Verification Status", "").strip() == "verified":
                notes_bits.append("✅ EMAIL VERIFIED")
            if r.get("City", "").strip():
                notes_bits.append(r["City"].strip())
            if r.get("Mobile Number", "").strip():
                notes_bits.append("Ph: " + r["Mobile Number"].strip())
            add(r.get("Company", "Myntra"), name, r.get("Title", ""), r.get("Email", ""),
                r.get("LinkedIn", ""), "; ".join(notes_bits))
else:
    print("MISSING:", mpath)

with open("extra_contacts.csv", "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["Company", "Name", "Title", "Email", "LinkedIn", "Action"])
    w.writerows(out)

print(f"extra_contacts.csv: {len(out)} contacts")
